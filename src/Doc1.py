import self as self
from openpyxl import *


class doc1():
    a1 = []
    a2 = []
    a3 = []
    a4 = []
    wb: Workbook = Workbook()
    ws = wb.active
    wb.create_sheet("Mysheet")
    wb.save('Звіт1.xlsx')
    ws = wb["Mysheet"]
    wcell1 = ws.cell(1, 1)
    wcell1.value = "Семестр 1"
    wcell2 = ws.cell(2, 1)
    wcell2.value = "К-сть груп"
    wcell3 = ws.cell(3, 1)
    wcell3.value = a1[1]
    wcell4 = ws.cell(2, 2)
    wcell4.value = "К-cть підгруп"
    wcell5 = ws.cell(3, 2)
    wcell5.value = a1[2]
    wcell6 = ws.cell(2, 3)
    wcell6.value = "К-cть потоків"
    wcell7 = ws.cell(3, 3)
    wcell7.value = a1[3]
    wcell8 = ws.cell(2, 4)
    wcell8.value = "Всього кредитів ECTS"
    wcell9 = ws.cell(3, 4)
    wcell9.value = a1[4]
    wcell10 = ws.cell(2, 5)
    wcell10.value = "Чит. лекцій"
    wcell11 = ws.cell(3, 5)
    wcell11.value = a1[11]
    wcell12 = ws.cell(2, 6)
    wcell12.value = "Провед. лабор. занять"
    wcell13 = ws.cell(3, 6)
    wcell13.value = a1[12]
    wcell14 = ws.cell(2, 7)
    wcell14.value = "Провед. практ./ семінар. занять"
    wcell15 = ws.cell(3, 7)
    wcell15.value = a1[13]
    wcell16 = ws.cell(2, 8)
    wcell16.value = "Пров. консульт з нав. дисц. протягом семестру"
    wcell17 = ws.cell(3, 8)
    wcell17.value = a1[14]
    wcell18 = ws.cell(2, 9)
    wcell18.value = "Пров. екзам. консультацій"
    wcell19 = ws.cell(3, 9)
    wcell19.value = a1[15]
    wcell20 = ws.cell(2, 10)
    wcell20.value = "Керівництво і приймання КП/КР"
    wcell21 = ws.cell(3, 10)
    wcell21.value = a1[16]
    wcell22 = ws.cell(2, 11)
    wcell22.value = "Пров. заліку"
    wcell23 = ws.cell(3, 11)
    wcell23.value = a1[17]
    wcell24 = ws.cell(2, 12)
    wcell24.value = "Пров. сем. екз."
    wcell25 = ws.cell(3, 12)
    wcell25.value = a1[18]
    wcell26 = ws.cell(2, 13)
    wcell26.value = "Кер-тво, консульт., реце-ня ДП"
    wcell27 = ws.cell(3, 13)
    wcell27.value = a1[19]
    wcell28 = ws.cell(2, 14)
    wcell28.value = "Пров-ня захисту"
    wcell29 = ws.cell(3, 14)
    wcell29.value = a1[20]
    wcell30 = ws.cell(2, 15)
    wcell30.value = "Пров-ня захисту"
    wcell31 = ws.cell(3, 15)
    wcell31.value = a1[21]
    wcell32 = ws.cell(2, 16)
    wcell32.value = "Кваліф. Іспит"
    wcell33 = ws.cell(3, 16)
    wcell33.value = a1[22]
    wcell34 = ws.cell(2, 17)
    wcell34.value = "Кер-тво НДРС"
    wcell35 = ws.cell(3, 17)
    wcell35.value = a1[23]
    wcell36 = ws.cell(2, 18)
    wcell36.value = "Кер-тво аспірантами, здобувачами"
    wcell37 = ws.cell(3, 18)
    wcell37.value = a1[24]
    wcell38 = ws.cell(2, 19)
    wcell38.value = "Кер-тво аспірантами, здобувачами"
    wcell39 = ws.cell(3, 19)
    wcell39.value = a1[25]
    wcell40 = ws.cell(2, 20)
    wcell40.value = "Кер-тво практ."
    wcell41 = ws.cell(3, 20)
    wcell41.value = a1[26]
    wcell42 = ws.cell(2, 21)
    wcell42.value = "Інші види -5%"
    wcell43 = ws.cell(3, 21)
    wcell43.value = a1[27]
    wcell44 = ws.cell(2, 22)
    wcell44.value = "Дист. Модуль"
    wcell45 = ws.cell(3, 22)
    wcell45.value = a1[28]
    wcell46 = ws.cell(2, 23)
    wcell46.value = "Додаткові години"
    wcell47 = ws.cell(3, 23)
    wcell47.value = a1[29]
    wcell48 = ws.cell(2, 24)
    wcell48.value = "Додаткові години"
    wcell49 = ws.cell(3, 24)
    wcell49.value = a1[30]
    wcell50 = ws.cell(2, 25)
    wcell50.value = "Всього"
    wcell51 = ws.cell(3, 25)
    wcell51.value = a1[31]
    wb.save("resources\\ІТтаКБ. Сем I. Форма навчання  денна.xlsx")

    wb.close()